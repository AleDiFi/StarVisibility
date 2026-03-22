"""
Excel formatter for StarVisibility (optional, requires openpyxl).

Produces a single .xlsx file with two sheets:
  - "Targets"  — full target list
  - "Summary"  — slot/sector coverage overview

Column headers are bold, columns are auto-sized, and alternating row
colours improve readability in Excel.

If openpyxl is not installed, calling any function in this module raises
ImportError with a clear installation instruction.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional

from src.config.settings import EXCEL_SHEET_SUMMARY, EXCEL_SHEET_TARGETS, OUTPUT_DIR
from src.io.csv_exporter import SUMMARY_COLUMNS, TARGET_COLUMNS
from src.models.domain import PlanningResult, SelectedTarget, SlotSectorCoverage
from src.utils.logging_utils import get_logger

log = get_logger("excel_formatter")


def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export.\n"
            "Install it with: pip install openpyxl"
        ) from exc


def export_to_excel(
    result: PlanningResult,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Export PlanningResult to a formatted Excel workbook.

    Parameters
    ----------
    result : PlanningResult
    output_path : explicit path; if None a timestamped file is placed in OUTPUT_DIR

    Returns
    -------
    Path of the written .xlsx file
    """
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"starvisibility_{ts}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ---- Targets sheet ----
    ws_targets = wb.active
    ws_targets.title = EXCEL_SHEET_TARGETS
    _write_sheet(
        ws_targets,
        columns=TARGET_COLUMNS,
        rows=[t.to_export_dict() for t in sorted(
            result.selected_targets,
            key=lambda t: (t.slot.night_label, t.slot.slot_index,
                           t.sector.name, t.mag_bin.label)
        )],
    )

    # ---- Summary sheet ----
    ws_summary = wb.create_sheet(EXCEL_SHEET_SUMMARY)
    _write_sheet(
        ws_summary,
        columns=SUMMARY_COLUMNS,
        rows=[c.to_export_dict() for c in sorted(
            result.coverage,
            key=lambda c: (c.night_label, c.slot_label, c.sector_name)
        )],
    )

    wb.save(output_path)
    log.info("Excel workbook written: %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

HEADER_FILL = "4472C4"    # blue
ALT_ROW_FILL = "DCE6F1"   # light blue


def _write_sheet(ws, columns: List[str], rows: List[dict]) -> None:
    """Write header + data rows to a worksheet with basic formatting."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        alt_fill = PatternFill("solid", fgColor=ALT_ROW_FILL)
        center = Alignment(horizontal="center", vertical="center")

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data rows
        for row_idx, row_data in enumerate(rows, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill

        # Auto-size columns (approximation)
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(col_name) + 2
            for row_idx in range(2, len(rows) + 2):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v:
                    max_length = max(max_length, len(str(v)) + 2)
            ws.column_dimensions[col_letter].width = min(max_length, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

    except ImportError:
        # openpyxl styling may differ between versions; write plain data
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx,
                        value=row_data.get(col_name, ""))
